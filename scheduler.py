import pandas as pd
from ortools.sat.python import cp_model
import holidays
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os

class WebSolutionCounter(cp_model.CpSolverSolutionCallback):
    def __init__(self, objectives, weights, queue_callback):
        super().__init__()
        self._solution_count = 0
        self._objectives = objectives
        self._weights = weights
        self.queue_callback = queue_callback
        self._display_order = [
            ('total_used_points', '總使用點數'), ('linear_gaps_bonus', '線性間隔獎勵'),
            ('min_gap_count', '隔兩天次數(懲罰)'), ('fairness_penalty', '同儕公平性(懲罰)'),
            ('total_shifts_filled', '總排班數量'), ('i_priority_bonus', 'I 區優先獎勵'),
            ('home_area_bonus', '在家區域獎勵')
        ]

    def on_solution_callback(self):
        self._solution_count += 1
        self.queue_callback(f"--- 找到第 {self._solution_count} 個可行解 ---")
        total_score = 0
        for key, display_name in self._display_order:
            if key in self._objectives:
                raw_val = self.Value(self._objectives[key])
                score = raw_val * self._weights[key]
                total_score += score
                self.queue_callback(f"  - {display_name:<12}: {raw_val:>5} (分數: {int(score)})")
        self.queue_callback(f"  >> 此解總分: {int(total_score)}")
        self.queue_callback("")

    def solution_count(self):
        return self._solution_count

def format_excel(writer, doctor_schedule_df, weekend_days, official_holidays, doctor_info):
    workbook = writer.book
    ws_doctor = writer.sheets['醫師月曆班表']
    colors = {'A': 'ADD8E6', 'B': '90EE90', 'C': 'FFFFE0', 'I': 'FFB6C1'}
    fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in colors.items()}
    weekend_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type="solid")
    holiday_fill = PatternFill(start_color='FFDDC1', end_color='FFDDC1', fill_type="solid")
    unavailable_fill = PatternFill(patternType='gray0625', fgColor='A9A9A9')
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type="solid")
    header_font = Font(bold=True, color='000000')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, day in enumerate(doctor_schedule_df.columns, 2):
        col_letter = get_column_letter(col_idx)
        fill_to_apply = None
        if day in official_holidays: fill_to_apply = holiday_fill
        elif day in weekend_days: fill_to_apply = weekend_fill
        if fill_to_apply:
            for row_idx in range(1, ws_doctor.max_row + 1):
                ws_doctor.cell(row=row_idx, column=col_idx).fill = fill_to_apply

    for r_idx, doc in enumerate(doctor_schedule_df.index, 2):
        ws_doctor.cell(row=r_idx, column=1).font = Font(bold=True)
        ws_doctor.row_dimensions[r_idx].height = 25
        for c_idx, day in enumerate(doctor_schedule_df.columns, 2):
            cell = ws_doctor.cell(row=r_idx, column=c_idx)
            cell.alignment = center_align
            cell.border = thin_border
            if cell.value and str(cell.value) in fills:
                cell.fill = fills[str(cell.value)]
    
    doc_to_row_map = {doc: i + 2 for i, doc in enumerate(doctor_schedule_df.index)}
    for doc, info in doctor_info.items():
        row_idx = doc_to_row_map[doc]
        for day_off in info['不可排班日']:
            if day_off in doctor_schedule_df.columns:
                day_col_idx = list(doctor_schedule_df.columns).index(day_off) + 2
                cell = ws_doctor.cell(row=row_idx, column=day_col_idx)
                cell.fill = unavailable_fill
                cell.value = "預休"
                cell.font = Font(color='FFFFFF', bold=True)

    for col_idx in range(1, ws_doctor.max_column + 1):
        cell = ws_doctor.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
        ws_doctor.column_dimensions[get_column_letter(col_idx)].width = 6
    ws_doctor.row_dimensions[1].height = 20
    ws_doctor.column_dimensions['A'].width = 12

    ws_summary = writer.sheets['點數統計總覽']
    for col_idx in range(1, ws_summary.max_column + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 15
        ws_summary.cell(row=1, column=col_idx).font = header_font
        ws_summary.cell(row=1, column=col_idx).fill = header_fill

# 修改函式簽名，接收 output_base_dir
def solve_schedule_web(doctor_data, year, month, q, output_base_dir):
    try:
        q.put("1. 正在讀取前端傳來的醫師資料...")
        df = pd.DataFrame(doctor_data)
        
        YEAR, MONTH = year, month
        num_days = pd.Period(f'{YEAR}-{MONTH}-01').days_in_month
        date_range = pd.to_datetime(pd.Series(pd.date_range(start=f'{YEAR}-{MONTH}-01', end=f'{YEAR}-{MONTH}-{num_days}')))
        weekend_days = date_range[date_range.dt.dayofweek.isin([5, 6])].dt.day.tolist()
        official_holidays = [d.day for d in holidays.TW(years=YEAR) if d.month == MONTH]
        double_point_days = set(weekend_days + official_holidays)
        doctors = df['醫師姓名'].tolist()
        areas = ['A', 'B', 'C', 'I']
        doctor_info = df.set_index('醫師姓名').to_dict('index')

        q.put("2. 正在建立數學模型...")
        model = cp_model.CpModel()
        
        shifts = {}
        for doc in doctors:
            for day in range(1, num_days + 1):
                for area in areas:
                    shifts[(doc, day, area)] = model.NewBoolVar(f'shift_{doc}_{day}_{area}')

        # --- 硬性規則 (與原始碼相同) ---
        for day in range(1, num_days + 1):
            for area in areas: model.AddAtMostOne([shifts[(doc, day, area)] for doc in doctors])
        for day in range(1, num_days + 1):
            for doc in doctors: model.AddAtMostOne([shifts[(doc, day, area)] for area in areas])
        i_doctors = [d for d, info in doctor_info.items() if info['區域'] == 'I']
        support_doctors = [d for d, info in doctor_info.items() if info['區域'] in ['A', 'B', 'C']]
        for day in range(1, num_days + 1): model.Add(sum(shifts[(doc, day, 'I')] for doc in support_doctors) == 0)
        for doc in doctors:
            for day in range(1, num_days - 1): model.Add(sum(shifts[doc, d, area] for area in areas for d in range(day, day+3)) <= 1)
        for doc in doctors:
            for day in doctor_info[doc]['不可排班日']:
                if 1 <= day <= num_days:
                    for area in areas: model.Add(shifts[(doc, day, area)] == 0)
        points_per_doctor = {}
        for doc in doctors:
            points_for_doc = sum(shifts[(doc, day, area)] * (2 if day in double_point_days else 1) for day in range(1, num_days + 1) for area in areas)
            points_per_doctor[doc] = points_for_doc
            model.Add(points_per_doctor[doc] <= doctor_info[doc]['點數上限'])

        # --- 軟性目標 (與原始碼相同) ---
        is_work_day = {}
        for doc in doctors:
            for day in range(1, num_days + 1):
                is_work_day[doc, day] = model.NewBoolVar(f'is_work_day_{doc}_{day}')
                model.Add(is_work_day[doc, day] == sum(shifts[doc, day, area] for area in areas))
        total_used_points = sum(points_per_doctor.values())
        all_linear_bonuses = []
        for doc in doctors:
            for d1 in range(1, num_days + 1):
                for d2 in range(d1 + 1, num_days + 1):
                    is_consecutive = model.NewBoolVar(f'consecutive_{doc}_{d1}_{d2}')
                    no_work_in_between_literals = [is_work_day[doc, d].Not() for d in range(d1 + 1, d2)]
                    model.AddBoolAnd([is_work_day[doc, d1], is_work_day[doc, d2]] + no_work_in_between_literals).OnlyEnforceIf(is_consecutive)
                    model.AddBoolOr([is_work_day[doc, d1].Not(), is_work_day[doc, d2].Not()] + [is_work_day[doc, d] for d in range(d1 + 1, d2)]).OnlyEnforceIf(is_consecutive.Not())
                    gap = d2 - d1
                    linear_bonus = 10 * gap
                    all_linear_bonuses.append(is_consecutive * linear_bonus)
        total_linear_gaps_bonus = sum(all_linear_bonuses)
        min_gap_penalties = []
        for doc in doctors:
            for day in range(1, num_days - 2):
                has_min_gap = model.NewBoolVar(f'has_min_gap_{doc}_{day}')
                model.AddBoolAnd([is_work_day[doc, day], is_work_day[doc, day + 3]]).OnlyEnforceIf(has_min_gap)
                model.AddBoolOr([is_work_day[doc, day].Not(), is_work_day[doc, day + 3].Not()]).OnlyEnforceIf(has_min_gap.Not())
                min_gap_penalties.append(has_min_gap)
        total_min_gap_count = sum(min_gap_penalties)
        peer_groups = defaultdict(list)
        for doc, info in doctor_info.items():
            key = (info['區域'], info['點數上限'])
            peer_groups[key].append(doc)
        all_ranges = []
        for group_key, group_docs in peer_groups.items():
            if len(group_docs) > 1:
                group_points = [points_per_doctor[doc] for doc in group_docs]
                min_points, max_points = model.NewIntVar(0, 100, ''), model.NewIntVar(0, 100, '')
                model.AddMinEquality(min_points, group_points)
                model.AddMaxEquality(max_points, group_points)
                group_range = model.NewIntVar(0, 100, '')
                model.Add(group_range == max_points - min_points)
                all_ranges.append(group_range)
        fairness_penalty = sum(all_ranges)
        total_shifts_filled = sum(shifts.values())
        i_priority_bonus = sum(shifts[(doc, day, 'I')] for doc in i_doctors for day in range(1, num_days + 1))
        home_area_bonus = sum(shifts[(doc, day, info['區域'])] for doc, info in doctor_info.items() for day in range(1, num_days + 1))
        
        objectives = {'total_used_points': total_used_points, 'linear_gaps_bonus': total_linear_gaps_bonus, 'min_gap_count': total_min_gap_count, 'fairness_penalty': fairness_penalty, 'total_shifts_filled': total_shifts_filled, 'i_priority_bonus': i_priority_bonus, 'home_area_bonus': home_area_bonus}
        weights = {'total_used_points': 10000, 'linear_gaps_bonus': 10, 'min_gap_count': -500, 'fairness_penalty': -200, 'total_shifts_filled': 100, 'i_priority_bonus': 10, 'home_area_bonus': 0.1}
        model.Maximize(sum(objectives[name] * weights[name] for name in objectives))

        solver = cp_model.CpSolver()
        solver.parameters.enumerate_all_solutions = True
        solution_counter = WebSolutionCounter(objectives, weights, q.put)
        solver.parameters.max_time_in_seconds = 120.0
        q.put(f"3. 正在運算，尋找所有可能的排班方案 ({YEAR}-{MONTH})...")
        status = solver.Solve(model, solution_counter)

        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            q.put("4. 已找到最佳解！正在生成視覺化報告...")
            
            final_schedule_data = defaultdict(dict)
            for doc in doctors:
                for day in range(1, num_days + 1):
                    for area in areas:
                        if solver.Value(shifts[(doc, day, area)]) == 1:
                            final_schedule_data[doc][day] = area

            points_summary_data = []
            for doc in doctors:
                points_used, days_worked = 0, []
                for day in range(1, num_days + 1):
                    if day in final_schedule_data.get(doc, {}):
                        area = final_schedule_data[doc][day]
                        points_used += 2 if day in double_point_days else 1
                        days_worked.append(f"{day}({area})")
                points_summary_data.append({'醫師姓名': doc, '區域': doctor_info[doc]['區域'], '點數上限': doctor_info[doc]['點數上限'], '實際點數': points_used, '剩餘點數': doctor_info[doc]['點數上限'] - points_used, '排班日與區域': ", ".join(days_worked)})
            
            final_scores = {display_name: solver.Value(objectives[key]) for key, display_name in solution_counter._display_order if key in objectives}

            doctor_schedule_df = pd.DataFrame('', index=doctors, columns=range(1, num_days + 1))
            for doc, day_map in final_schedule_data.items():
                for day, area in day_map.items():
                    if doc in doctor_schedule_df.index and day in doctor_schedule_df.columns:
                        doctor_schedule_df.loc[doc, day] = area
            
            points_summary_df = pd.DataFrame(points_summary_data)
            
            schedule_df = pd.DataFrame("", index=range(1, num_days + 1), columns=areas)
            for doc, day_area_map in final_schedule_data.items():
                for day, area in day_area_map.items():
                    schedule_df.loc[day, area] = doc

            output_filename = f'schedule_result_{YEAR}-{MONTH}.xlsx'
            # 使用傳入的 output_base_dir 來組合正確的檔案路徑
            output_filepath = os.path.join(output_base_dir, output_filename)
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                doctor_schedule_df.to_excel(writer, sheet_name='醫師月曆班表')
                points_summary_df.to_excel(writer, sheet_name='點數統計總覽', index=False)
                schedule_df.to_excel(writer, sheet_name='區域班表')
                format_excel(writer, doctor_schedule_df, weekend_days, official_holidays, doctor_info)

            q.put(f"\n✅ **排班完成！**")
            q.put(f"   詳細結果請點擊下方的按鈕下載 **{output_filename}** 檔案。")
            q.put("\n--- 最終排班結果分析 ---")
            q.put(f"在所有規則限制下，系統總共找到了 **{solution_counter.solution_count()}** 種不同的可行排班方案。")
            q.put(f"呈現的是其中一個綜合評分最高的「最佳解」。")
            
            result_payload = {
                "status": "success", "final_scores": final_scores,
                "schedule_data": {
                    "doctors": doctors, "num_days": num_days,
                    "schedule": final_schedule_data,
                    "holidays": official_holidays,
                    "weekends": weekend_days,
                    "days_off": {doc: info['不可排班日'] for doc, info in doctor_info.items()},
                    "doctor_info": doctor_info
                },
                "points_summary_html": points_summary_df.to_html(classes='table table-hover', index=False),
                "area_schedule_html": schedule_df.to_html(classes='table table-bordered table-sm text-center'),
                "excel_url": f"/output/{output_filename}"
            }
            q.put(result_payload)
        else:
            q.put("\n❌ **錯誤：** 在目前的規則下，找不到任何可行的排班解。")
            q.put({"status": "error", "message": "找不到可行的排班解。"})
    
    except Exception as e:
        q.put(f"\n❌ **後端錯誤：** {e}")
        q.put({"status": "error", "message": f"後端發生錯誤: {e}"})
    finally:
        q.put("DONE")